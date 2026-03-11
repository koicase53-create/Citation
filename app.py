# --- START OF FILE app.py (CORRECTED FOR EXCEL EXPORT) ---

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash, Response
import logging, os, json, sqlite3, io, threading, re, glob
from collections import defaultdict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
from urllib.parse import quote 

from journal_catalog import (
    classify_journal,
    get_catalog_path,
    get_frontend_categories,
    get_journal_name_variants,
    load_catalog,
    normalize_journal_name,
)


app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-only-secret-key-change-in-production")
CORS(app)
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


@app.context_processor
def inject_template_globals():
    try:
        script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "script.js")
        asset_version = int(os.path.getmtime(script_path))
    except Exception:
        asset_version = int(datetime.now().timestamp())
    return {
        "current_year": datetime.now().year,
        "asset_version": asset_version,
    }

# --- 配置 ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_DATABASE_FILE = os.path.join(BASE_DIR, 'citations.db')
OPTIMIZED_DATABASE_FILE = os.path.join(BASE_DIR, 'citations_optimized.db')
DATABASE_FILE = os.environ.get(
    "CITATIONS_DB_FILE",
    OPTIMIZED_DATABASE_FILE if os.path.exists(OPTIMIZED_DATABASE_FILE) else DEFAULT_DATABASE_FILE,
)
DATABASE_FILES_ENV = os.environ.get("CITATIONS_DB_FILES", "").strip()


def _resolve_db_path(path):
    if not path:
        return ""
    resolved = os.path.expanduser(path.strip())
    if not os.path.isabs(resolved):
        resolved = os.path.join(BASE_DIR, resolved)
    return os.path.abspath(resolved)


if DATABASE_FILES_ENV:
    DATABASE_FILES = [_resolve_db_path(p) for p in DATABASE_FILES_ENV.split(",") if p.strip()]
else:
    auto_shards = sorted(glob.glob(os.path.join(BASE_DIR, "citations_shard_*.db")))
    if len(auto_shards) > 1:
        DATABASE_FILES = [_resolve_db_path(p) for p in auto_shards]
    else:
        DATABASE_FILES = [_resolve_db_path(DATABASE_FILE)]

DATABASE_FILE = DATABASE_FILES[0] if DATABASE_FILES else _resolve_db_path(DATABASE_FILE)
IS_SHARDED_DATABASE = len(DATABASE_FILES) > 1
SHARD_ALIASES = [f"db{i}" for i in range(len(DATABASE_FILES))]

USERS_FILE = os.path.join(BASE_DIR, 'users.json')
INVITATION_CODE = '2025888'
SEARCH_LOG_FILE = os.path.join(BASE_DIR, 'search_log.log')
log_lock = threading.Lock()

CATALOG_PATH = get_catalog_path(BASE_DIR)
JOURNAL_CATALOG = load_catalog(CATALOG_PATH)
LATEST_LEVELS_MAP = JOURNAL_CATALOG.get("latest_levels", {})
FRONTEND_JOURNAL_CATEGORIES = get_frontend_categories(JOURNAL_CATALOG)
if not LATEST_LEVELS_MAP:
    logging.warning("未加载到最新期刊目录，将仅使用旧版分类逻辑。目录路径: %s", CATALOG_PATH)
if IS_SHARDED_DATABASE:
    logging.info("已启用多库分片模式，共 %d 个数据库文件。", len(DATABASE_FILES))

_DB_JOURNAL_LOOKUP = None
_FALLBACK_FRONTEND_JOURNAL_CATEGORIES = None


def _db_file_has_table(db_path, table_name):
    if not os.path.exists(db_path):
        return False
    conn = sqlite3.connect(db_path)
    try:
        row = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,),
        ).fetchone()
        return bool(row)
    finally:
        conn.close()


def build_fallback_frontend_categories():
    """
    当最新目录 xlsx 缺失或解析失败时，
    基于数据库中的期刊字段构建前端可用分类，避免快捷按钮失效。
    """
    global _FALLBACK_FRONTEND_JOURNAL_CATEGORIES
    if _FALLBACK_FRONTEND_JOURNAL_CATEGORIES is not None:
        return _FALLBACK_FRONTEND_JOURNAL_CATEGORIES

    level_label_map = {
        "CSSCI": "CSSCI 来源期刊(C刊)",
        "C扩": "CSSCI 扩展版来源期刊(C扩)",
        "集刊": "CSSCI 集刊",
    }
    level_priority = {"CSSCI": 3, "C扩": 2, "集刊": 1}
    level_to_journals = {level: set() for level in level_label_map}

    conn = get_db_connection()
    try:
        rows = conn.execute(
            """
            SELECT Journal_Name, MAX(Year) AS latest_year
            FROM articles
            WHERE Journal_Name IS NOT NULL AND TRIM(Journal_Name) != ''
            GROUP BY Journal_Name
            """
        ).fetchall()

        journal_best_level = {}
        for row in rows:
            journal_name = row["Journal_Name"]
            latest_year = row["latest_year"] or 0
            level = get_journal_level(journal_name, latest_year)
            if level not in level_label_map:
                continue

            existing_level = journal_best_level.get(journal_name)
            if existing_level is None or level_priority[level] > level_priority[existing_level]:
                journal_best_level[journal_name] = level

        for journal_name, level in journal_best_level.items():
            level_to_journals[level].add(journal_name)
    except Exception as exc:
        logging.warning("构建前端期刊分类兜底数据失败: %s", exc)
    finally:
        conn.close()

    result = {}
    for level, label in level_label_map.items():
        journals = sorted(level_to_journals[level])
        if journals:
            result[label] = {
                "subcategories": {
                    "默认分组": [{"cnTitle": name} for name in journals]
                }
            }

    _FALLBACK_FRONTEND_JOURNAL_CATEGORIES = result
    return result


def get_journal_type(journal_name, year):
    _, journal_type, _ = classify_journal(journal_name, year, LATEST_LEVELS_MAP)
    return journal_type


def get_journal_level(journal_name, year):
    journal_level, _, _ = classify_journal(journal_name, year, LATEST_LEVELS_MAP)
    return journal_level


def build_journal_fts_query(journal_names_list):
    """
    为期刊列表构建 FTS5 查询表达式
    使用基础名称（去掉括号）来匹配，避免FTS5对括号的处理问题
    """
    all_variants = []
    
    for journal_name in journal_names_list:
        if not journal_name or not journal_name.strip():
            continue
        
        # 获取该期刊的所有变体（基础名称）
        variants = get_journal_name_variants(journal_name.strip())
        all_variants.extend(variants)
    
    # 去重
    unique_variants = list(set(v for v in all_variants if v))
    
    if not unique_variants:
        return '""'  # 返回一个不会匹配任何内容的查询
    
    # 构建 FTS 表达式：用 OR 连接所有变体
    # 使用引号确保短语匹配，避免分词问题
    quoted_variants = [f'"{variant}"' for variant in unique_variants]
    
    return ' OR '.join(quoted_variants)


def find_matching_journals_in_db(conn, user_input_journals):
    """
    在数据库中查找与用户输入匹配的实际期刊名称
    返回数据库中的实际期刊名称列表
    """
    global _DB_JOURNAL_LOOKUP

    if _DB_JOURNAL_LOOKUP is None:
        lookup = defaultdict(set)
        rows = conn.execute(
            "SELECT DISTINCT Journal_Name FROM articles WHERE Journal_Name IS NOT NULL"
        ).fetchall()
        for row in rows:
            journal_name = row["Journal_Name"]
            normalized_name = normalize_journal_name(journal_name)
            if normalized_name:
                lookup[normalized_name].add(journal_name)
        _DB_JOURNAL_LOOKUP = lookup

    matched_journals = set()
    for user_journal in user_input_journals:
        normalized_user = normalize_journal_name(user_journal.strip())
        if normalized_user and normalized_user in _DB_JOURNAL_LOOKUP:
            matched_journals.update(_DB_JOURNAL_LOOKUP[normalized_user])

    return sorted(matched_journals)


def has_article_journal_links_table():
    return any(_db_file_has_table(db_path, "article_journal_links") for db_path in DATABASE_FILES)


def ensure_runtime_indexes():
    for db_path in DATABASE_FILES:
        if not os.path.exists(db_path):
            continue

        conn = sqlite3.connect(db_path)
        try:
            conn.executescript(
                """
                CREATE INDEX IF NOT EXISTS idx_articles_article_id ON articles(Article_ID);
                CREATE INDEX IF NOT EXISTS idx_articles_year ON articles(Year);
                CREATE INDEX IF NOT EXISTS idx_articles_journal_name ON articles(Journal_Name);
                CREATE INDEX IF NOT EXISTS idx_articles_doi ON articles(DOI);
                """
            )
            conn.commit()
        except Exception as exc:
            logging.warning("创建运行时索引失败 (%s): %s", db_path, exc)
        finally:
            conn.close()


ensure_runtime_indexes()


# --- 异常处理 ---
class ApiException(Exception):
    def __init__(self, message, status_code=400):
        super().__init__(message)
        self.message = message
        self.status_code = status_code

@app.errorhandler(ApiException)
def handle_api_exception(error):
    return jsonify({"success": False, "error": error.message}), error.status_code

@app.errorhandler(Exception)
def handle_generic_exception(error):
    if isinstance(error, HTTPException):
        return error
    logging.error(f"发生未捕获的异常: {error}", exc_info=True)
    return jsonify({"success": False, "error": f"服务器发生内部错误: {str(error)}"}), 500

# --- 日志与用户管理 ---
def log_search_activity(search_type, **kwargs):
    user = session.get('user_email', 'anonymous')
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    search_terms = ', '.join(f'{k}: {v}' for k, v in kwargs.items() if v)
    log_entry = f"{timestamp} | User: {user} | Type: {search_type} | Terms: {search_terms}\n"
    
    with log_lock:
        try:
            with open(SEARCH_LOG_FILE, 'a', encoding='utf-8') as f:
                f.write(log_entry)
        except Exception as e:
            logging.error(f"Failed to write to search log: {e}")

def load_users():
    if not os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'w') as f: json.dump({}, f)
        return {}
    with open(USERS_FILE, 'r') as f:
        try: return json.load(f)
        except json.JSONDecodeError: return {}

def save_user(email, password):
    users = load_users()
    users[email] = password
    with open(USERS_FILE, 'w') as f: json.dump(users, f, indent=4)

# --- 认证与页面路由 ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email, password = request.form.get('email'), request.form.get('password')
        users = load_users()
        if email in users and users[email] == password:
            session['user_email'] = email
            flash('登录成功！', 'success')
            return redirect(url_for('welcome'))
        else:
            flash('邮箱或密码错误。', 'error')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        invite_code, email, password = request.form.get('invite_code'), request.form.get('email'), request.form.get('password')
        if invite_code != INVITATION_CODE:
            flash('无效的邀请码。', 'error')
            return redirect(url_for('register'))
        users = load_users()
        if email in users:
            flash('该邮箱已被注册。', 'error')
            return redirect(url_for('register'))
        if not email or not password:
            flash('邮箱和密码不能为空。', 'error')
            return redirect(url_for('register'))
        save_user(email, password)
        session['user_email'] = email
        flash('注册成功！', 'success')
        return redirect(url_for('welcome'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.pop('user_email', None)
    flash('您已成功登出。', 'success')
    return redirect(url_for('login'))

@app.route('/')
def welcome():
    if 'user_email' not in session: return redirect(url_for('login'))
    return render_template('welcome.html')

@app.route('/tool')
def tool():
    if 'user_email' not in session: return redirect(url_for('login'))
    return render_template('index.html')

# --- 数据库辅助函数 ---
def _apply_connection_pragmas(conn):
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA temp_store = MEMORY")
    conn.execute("PRAGMA cache_size = -20000")
    conn.execute("PRAGMA mmap_size = 268435456")


def _attached_table_exists(conn, alias, table_name):
    row = conn.execute(
        f"SELECT 1 FROM {alias}.sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    ).fetchone()
    return bool(row)


def _create_union_view(conn, view_name, source_aliases, table_name, columns="*"):
    if not source_aliases:
        return
    union_sql = " UNION ALL ".join(
        [f"SELECT {columns} FROM {alias}.{table_name}" for alias in source_aliases]
    )
    conn.execute(f"CREATE TEMP VIEW {view_name} AS {union_sql}")


def _fts_query_rowids(conn, column_name, match_expression):
    if not match_expression:
        return set()

    allowed_columns = {"Title", "Authors", "Institution", "Reference", "Keywords"}
    if column_name not in allowed_columns:
        raise ApiException(f"不支持的全文检索字段: {column_name}", 500)

    rowids = set()
    if IS_SHARDED_DATABASE:
        for alias in SHARD_ALIASES:
            query = f"SELECT rowid FROM {alias}.articles_fts WHERE {column_name} MATCH ?"
            rows = conn.execute(query, (match_expression,)).fetchall()
            rowids.update(row["rowid"] for row in rows)
    else:
        query = f"SELECT rowid FROM articles_fts WHERE {column_name} MATCH ?"
        rows = conn.execute(query, (match_expression,)).fetchall()
        rowids.update(row["rowid"] for row in rows)
    return rowids


def get_db_connection():
    missing = [db_path for db_path in DATABASE_FILES if not os.path.exists(db_path)]
    if missing:
        raise ApiException(f"数据库文件未找到: {', '.join(missing)}", 500)

    if not IS_SHARDED_DATABASE:
        conn = sqlite3.connect(DATABASE_FILE)
        _apply_connection_pragmas(conn)
        return conn

    conn = sqlite3.connect(":memory:")
    _apply_connection_pragmas(conn)

    for alias, db_path in zip(SHARD_ALIASES, DATABASE_FILES):
        conn.execute(f"ATTACH DATABASE ? AS {alias}", (db_path,))

    _create_union_view(conn, "articles", SHARD_ALIASES, "articles")

    link_aliases = [a for a in SHARD_ALIASES if _attached_table_exists(conn, a, "article_journal_links")]
    if link_aliases:
        _create_union_view(conn, "article_journal_links", link_aliases, "article_journal_links")

    return conn

def _db_rows_to_articles_json(rows, dynamic_counts=None):
    articles = []
    for row in rows:
        row_dict = dict(row)
        article_id = str(row_dict.get('Article_ID', ''))
        citedby_count = dynamic_counts.get(article_id) if dynamic_counts is not None else row_dict.get('Citation_Frequency', 0)
        articles.append({
            "id": f"local:{article_id}",
            "doi": row_dict.get('DOI'),
            "title": row_dict.get('Title'),
            "publicationName": row_dict.get('Journal_Name'),
            "journal_type": row_dict.get('Journal_Type'),
            "coverDate": str(row_dict.get('Year')),
            "fullDate": row_dict.get('Date'),
            "url": f"https://doi.org/{row_dict.get('DOI')}" if row_dict.get('DOI') and row_dict.get('DOI') != '暂无' else '#',
            "authors": row_dict.get('Authors'),
            "institution": row_dict.get('Institution'),
            "citedby_count": int(citedby_count or 0),
            "abstract": row_dict.get('Abstract', '摘要不可用。'),
            "keywords": row_dict.get('Keywords'),
            "reference": row_dict.get('Reference')
        })
    return articles

def _extract_main_title(title):
    if not title or pd.isna(title):
        return ''
    return str(title).split('——')[0].split('：')[0].split(':')[0].strip()


def _normalize_match_text(text):
    normalized = str(text or '').lower()
    normalized = re.sub(r'\s+', '', normalized)
    normalized = re.sub(r'[“”"\'‘’`·•,，。:：;；!?！？()\[\]（）《》<>【】_/\\\-]', '', normalized)
    return normalized


def _split_authors(authors_text):
    if not authors_text:
        return []
    return [a.strip() for a in re.split(r'[;；]+', str(authors_text)) if a and a.strip()]


_SUPERSCRIPT_TO_DIGIT = str.maketrans("⁰¹²³⁴⁵⁶⁷⁸⁹", "0123456789")


def _normalize_affiliation_name(name):
    text = str(name or "").translate(_SUPERSCRIPT_TO_DIGIT).strip()
    if not text:
        return ""

    text = re.sub(r'^\s*[\(\[（]?\d+(?:\s*[,，、\-]\s*\d+)*[\)\]）]?\s*[\.．、:：]?\s*', '', text)
    text = text.replace("／", "/").replace("，", "、")
    text = re.sub(r'\s*/\s*', '/', text)
    text = re.sub(r'\s*、\s*', '、', text)
    text = re.sub(r'\s+', ' ', text).strip(" ;；,，")
    return text


def _affiliation_key(name):
    normalized = _normalize_affiliation_name(name)
    return re.sub(r'[\s,，;；:：。·•/、\-\(\)（）\[\]【】]', '', normalized)


def _parse_author_with_markers(author_raw):
    text = str(author_raw or "").translate(_SUPERSCRIPT_TO_DIGIT).strip()
    if not text:
        return "", set()

    marker_ids = set()
    bracket_match = re.search(r'[\(\[（]\s*([\d,\s，、-]+)\s*[\)\]）]\s*$', text)
    if bracket_match:
        marker_ids.update(int(v) for v in re.findall(r'\d+', bracket_match.group(1)))
        text = text[:bracket_match.start()].strip()

    tail_match = re.search(r'(\d+(?:\s*[,，、]\s*\d+)*)\s*$', text)
    if tail_match:
        marker_ids.update(int(v) for v in re.findall(r'\d+', tail_match.group(1)))
        text = text[:tail_match.start()].strip()

    cleaned_name = re.sub(r'[,\s]+$', '', text).strip()
    return cleaned_name, marker_ids


def _parse_institution_entries(institution_raw):
    entries = []
    if not institution_raw:
        return entries

    normalized_text = str(institution_raw).translate(_SUPERSCRIPT_TO_DIGIT).strip()
    if not normalized_text:
        return entries

    parts = []
    for part in re.split(r'[;；\n]+', normalized_text):
        chunk = str(part or "").strip()
        if not chunk:
            continue

        # CNKI 常见格式：1.xxx, 2.yyy（同一字段多机构）
        # 仅在逗号后紧跟编号时拆分，避免误拆普通机构名称。
        subparts = re.split(r'[，,、]\s*(?=[\(\[（]?\d+\s*[\.．、:：])', chunk)
        for sub in subparts:
            token = sub.strip(" ,，、")
            if token:
                parts.append(token)

    for token in parts:
        token = str(token or "").strip()
        if not token:
            continue

        marker_ids = set()
        match = re.match(r'^\s*[\(\[（]?(\d+(?:\s*[,，、\-]\s*\d+)*)[\)\]）]?\s*[\.．、:：]?\s*(.+)$', token)
        name_part = token
        if match:
            marker_ids.update(int(v) for v in re.findall(r'\d+', match.group(1)))
            name_part = match.group(2)

        affiliation_name = _normalize_affiliation_name(name_part)
        if affiliation_name:
            entries.append({
                "name": affiliation_name,
                "marker_ids": marker_ids,
            })
    return entries


def _aggregate_affiliations(affiliation_names):
    grouped = {}
    for name in affiliation_names:
        normalized_name = _normalize_affiliation_name(name)
        if not normalized_name:
            continue
        key = _affiliation_key(normalized_name)
        if not key:
            continue
        item = grouped.setdefault(key, {"name": normalized_name, "count": 0})
        item["count"] += 1
        if len(normalized_name) > len(item["name"]):
            item["name"] = normalized_name

    items = sorted(grouped.values(), key=lambda x: (-x["count"], x["name"]))
    return items


def _extract_author_affiliations(authors_raw, institutions_raw):
    authors = _split_authors(authors_raw)
    institution_entries = _parse_institution_entries(institutions_raw)

    marker_to_names = defaultdict(list)
    plain_institutions = []
    for entry in institution_entries:
        if entry["marker_ids"]:
            for marker_id in entry["marker_ids"]:
                marker_to_names[marker_id].append(entry["name"])
        else:
            plain_institutions.append(entry["name"])

    ordered_marker_institutions = []
    seen_ordered = set()
    for marker_id in sorted(marker_to_names.keys()):
        for name in marker_to_names[marker_id]:
            key = _affiliation_key(name)
            if key and key not in seen_ordered:
                ordered_marker_institutions.append(name)
                seen_ordered.add(key)

    result = {}
    for idx, author in enumerate(authors):
        author_name, marker_ids = _parse_author_with_markers(author)
        if not author_name:
            continue

        affiliations = []
        for marker_id in sorted(marker_ids):
            affiliations.extend(marker_to_names.get(marker_id, []))

        if not affiliations and plain_institutions and len(plain_institutions) == len(authors):
            affiliations = [plain_institutions[idx]]
        elif not affiliations and plain_institutions and len(plain_institutions) == 1:
            affiliations = [plain_institutions[0]]
        elif not affiliations and ordered_marker_institutions:
            # 常见脏数据：作者无编号、机构有编号。
            # 1) 若数量可一一对应，按顺序映射；
            # 2) 否则返回全部候选机构，避免误标为“未知机构”。
            if len(ordered_marker_institutions) == len(authors):
                affiliations = [ordered_marker_institutions[idx]]
            elif len(ordered_marker_institutions) == 1:
                affiliations = [ordered_marker_institutions[0]]
            else:
                affiliations = list(ordered_marker_institutions)

        if not affiliations:
            affiliations = ["未知机构"]

        merged_items = _aggregate_affiliations(affiliations)
        result[author_name] = [item["name"] for item in merged_items] or ["未知机构"]

    return result


def _parse_reference_entries(reference_text):
    text = str(reference_text or '').replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r'^\s*参考文献[:：]?\s*', '', text).strip()
    if not text:
        return []

    # Primary split: each [n] block is treated as one reference entry.
    chunks = [c.strip() for c in re.split(r'(?=\[\d+\])', text) if c.strip()]
    if len(chunks) > 1:
        return chunks

    # Fallback for references stored as plain lines without [n] delimiters.
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    if len(lines) <= 1:
        return lines

    entries = []
    current_entry = ''
    for line in lines:
        if re.match(r'^\[\d+\]', line):
            if current_entry:
                entries.append(current_entry.strip())
            current_entry = line
        else:
            current_entry = f"{current_entry} {line}".strip() if current_entry else line
    if current_entry:
        entries.append(current_entry.strip())
    return entries


def _build_target_signatures(target_rows):
    signatures = []
    seen = set()

    for row in target_rows:
        title = _extract_main_title(row['Title'] if isinstance(row, sqlite3.Row) else row.get('Title'))
        title_norm = _normalize_match_text(title)
        if not title_norm:
            continue

        authors_text = row['Authors'] if isinstance(row, sqlite3.Row) else row.get('Authors')
        author_norms = []
        for author in _split_authors(authors_text):
            norm_author = _normalize_match_text(author)
            if norm_author and norm_author not in author_norms:
                author_norms.append(norm_author)

        year_raw = row['Year'] if isinstance(row, sqlite3.Row) else row.get('Year')
        try:
            year = int(year_raw) if year_raw is not None and not pd.isna(year_raw) else None
        except (TypeError, ValueError):
            year = None

        journal_name = row['Journal_Name'] if isinstance(row, sqlite3.Row) else row.get('Journal_Name')
        journal_norms = []
        if journal_name and not pd.isna(journal_name):
            for journal_variant in get_journal_name_variants(str(journal_name)) + [str(journal_name)]:
                norm_journal = _normalize_match_text(journal_variant)
                if norm_journal and norm_journal not in journal_norms:
                    journal_norms.append(norm_journal)

        article_id = row['Article_ID'] if isinstance(row, sqlite3.Row) else row.get('Article_ID')
        dedupe_key = (str(article_id), title_norm, tuple(author_norms), year)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        signatures.append({
            "article_id": str(article_id),
            "title": title,
            "title_norm": title_norm,
            "title_is_generic": len(title_norm) <= 6,
            "author_norms": author_norms,
            "year": year,
            "journal_norms": journal_norms,
        })

    return signatures


def _reference_entry_matches_target(entry, target):
    entry_norm = _normalize_match_text(entry)
    if not entry_norm or target["title_norm"] not in entry_norm:
        return False

    author_matched = False
    if target["author_norms"]:
        author_matched = any(a in entry_norm for a in target["author_norms"])
        if not author_matched and target["title_is_generic"]:
            years_in_entry = set(re.findall(r'(?:19|20)\d{2}', entry))
            year_ok = (target["year"] is None) or (str(target["year"]) in years_in_entry)
            journal_ok = any(j in entry_norm for j in target["journal_norms"]) if target["journal_norms"] else False
            if not (year_ok and journal_ok):
                return False

    years_in_entry = set(re.findall(r'(?:19|20)\d{2}', entry))
    if target["year"] is not None and years_in_entry and str(target["year"]) not in years_in_entry:
        return False

    if target["title_is_generic"] and target["journal_norms"]:
        if not any(j in entry_norm for j in target["journal_norms"]):
            # Generic titles (e.g. “主持人语”) must match journal too.
            return False

    return True


def _get_citing_articles_for_titles(conn, target_titles, start_year, end_year):
    if not target_titles:
        return []

    quoted_titles = []
    for title in target_titles:
        main_title = _extract_main_title(title)
        if main_title:
            quoted_titles.append(f'"{main_title}"')

    if not quoted_titles:
        return []

    fts_expression = ' OR '.join(sorted(set(quoted_titles)))
    matching_ids = _fts_query_rowids(conn, "Reference", fts_expression)
    return _fetch_articles_by_ids(conn, matching_ids, start_year=start_year, end_year=end_year)


def _get_citing_articles_for_targets(conn, target_rows, start_year, end_year):
    signatures = _build_target_signatures(target_rows)
    if not signatures:
        return []

    coarse_rows = _get_citing_articles_for_titles(conn, [s["title"] for s in signatures], start_year, end_year)
    if not coarse_rows:
        return []

    filtered_rows = []
    seen_article_ids = set()
    for row in coarse_rows:
        ref_text = row['Reference'] or ''
        ref_entries = _parse_reference_entries(ref_text)
        matched = any(
            _reference_entry_matches_target(entry, signature)
            for entry in ref_entries
            for signature in signatures
        )
        if matched and row['Article_ID'] not in seen_article_ids:
            filtered_rows.append(row)
            seen_article_ids.add(row['Article_ID'])
    return filtered_rows


def _count_citations_for_targets(conn, target_rows, start_year, end_year):
    signatures = _build_target_signatures(target_rows)
    if not signatures:
        return {}

    coarse_rows = _get_citing_articles_for_titles(conn, [s["title"] for s in signatures], start_year, end_year)
    if not coarse_rows:
        return {s["article_id"]: 0 for s in signatures}

    parsed_entries_cache = {
        row['Article_ID']: _parse_reference_entries(row['Reference'] or '')
        for row in coarse_rows
    }

    citation_counts = {}
    for signature in signatures:
        matched_count = 0
        for row in coarse_rows:
            entries = parsed_entries_cache.get(row['Article_ID'], [])
            if any(_reference_entry_matches_target(entry, signature) for entry in entries):
                matched_count += 1
        citation_counts[signature["article_id"]] = matched_count
    return citation_counts


def _add_dynamic_citations_to_results(conn, article_rows, start_year, end_year):
    if not article_rows:
        return []
    dynamic_counts = _count_citations_for_targets(conn, article_rows, start_year, end_year)
    return _db_rows_to_articles_json(article_rows, dynamic_counts=dynamic_counts)


SQLITE_PARAM_LIMIT = 900


def _chunk_values(values, chunk_size=SQLITE_PARAM_LIMIT):
    values = list(values)
    for i in range(0, len(values), chunk_size):
        yield values[i:i + chunk_size]


def _normalize_target_journals(target_journals_list):
    normalized = {}
    for journal in target_journals_list:
        norm_name = normalize_journal_name(journal)
        if norm_name and norm_name not in normalized:
            normalized[norm_name] = journal
    return normalized


def _fetch_articles_by_ids(conn, article_ids, start_year=None, end_year=None):
    if not article_ids:
        return []

    rows = []
    for chunk in _chunk_values(sorted(article_ids)):
        placeholders = ",".join(["?"] * len(chunk))
        query = f"SELECT * FROM articles WHERE Article_ID IN ({placeholders})"
        params = list(chunk)
        if start_year is not None and end_year is not None:
            query += " AND Year >= ? AND Year <= ?"
            params.extend([int(start_year), int(end_year)])
        rows.extend(conn.execute(query, tuple(params)).fetchall())
    return rows


def _fetch_journal_links(conn, article_ids, target_norms=None):
    if not article_ids:
        return defaultdict(set)

    links_map = defaultdict(set)
    target_norms = sorted(set(target_norms or []))

    for chunk in _chunk_values(sorted(article_ids)):
        placeholders = ",".join(["?"] * len(chunk))
        query = (
            f"SELECT article_id, journal_normalized "
            f"FROM article_journal_links WHERE article_id IN ({placeholders})"
        )
        params = list(chunk)
        if target_norms:
            target_placeholders = ",".join(["?"] * len(target_norms))
            query += f" AND journal_normalized IN ({target_placeholders})"
            params.extend(target_norms)

        for row in conn.execute(query, tuple(params)).fetchall():
            links_map[row["article_id"]].add(row["journal_normalized"])

    return links_map


# --- API Endpoints ---
@app.route('/api/get-journal-catalog', methods=['POST'])
def get_journal_catalog_endpoint():
    catalog = FRONTEND_JOURNAL_CATEGORIES
    if not catalog:
        catalog = build_fallback_frontend_categories()
    return jsonify({
        "success": True,
        "catalog": catalog
    })


@app.route('/api/search-author-by-name', methods=['POST'])
def search_author_by_name_endpoint():
    data = request.get_json()
    query_term = data.get('name', '').strip()
    log_search_activity('查作者或机构', query=query_term)
    if not query_term:
        raise ApiException("作者姓名或机构名是必填项。")
    conn = get_db_connection()
    safe_term = query_term.replace('"', '').strip()
    prefix_query = f"{safe_term}*"
    author_ids = _fts_query_rowids(conn, "Authors", prefix_query)
    institution_ids = _fts_query_rowids(conn, "Institution", prefix_query)
    matched_ids = author_ids.union(institution_ids)
    matched_rows = _fetch_articles_by_ids(conn, matched_ids)
    
    author_works = defaultdict(list)
    for row in matched_rows:
        author_affiliation_map = _extract_author_affiliations(row['Authors'], row['Institution'])
        for author_name, institution_list in author_affiliation_map.items():
            if author_name.startswith(query_term):
                author_works[author_name].append({
                    'Article_ID': row['Article_ID'],
                    'Title': row['Title'],
                    'Authors': row['Authors'],
                    'Year': row['Year'],
                    'Journal_Name': row['Journal_Name'],
                    'institutions': institution_list,
                })

    if not author_works:
        conn.close()
        return jsonify({"success": True, "authors": []})

    all_target_rows = [work for works_list in author_works.values() for work in works_list]
    citation_counts_map = _count_citations_for_targets(conn, all_target_rows, 1900, 9999)

    authors_details = []
    for author_name, works in author_works.items():
        raw_affiliations = []
        for work in works:
            institutions = work.get('institutions') or []
            if not institutions:
                institutions = ["未知机构"]
            raw_affiliations.extend(institutions)

        affiliation_items = _aggregate_affiliations(raw_affiliations)
        if not affiliation_items:
            affiliation_items = [{"name": "未知机构", "count": 1}]

        cited_by_count = sum(citation_counts_map.get(str(work['Article_ID']), 0) for work in works)
            
        authors_details.append({
            "author_id": author_name,
            "name": author_name,
            "affiliation": '、'.join(item["name"] for item in affiliation_items),
            "affiliations": affiliation_items,
            "works_count": len(works),
            "cited_by_count": cited_by_count
        })

    conn.close()
    return jsonify({"success": True, "authors": sorted(authors_details, key=lambda x: x['cited_by_count'], reverse=True)})

@app.route('/api/get-works-with-dynamic-citations', methods=['POST'])
def get_works_with_dynamic_citations_endpoint():
    data = request.get_json()
    author_name = data.get('author_id')
    start_year = int(data.get('start_year', 1900))
    end_year = int(data.get('end_year', 9999))

    if not author_name:
        raise ApiException("未提供作者ID。")
        
    conn = get_db_connection()
    try:
        author_ids = _fts_query_rowids(conn, "Authors", f'"{author_name}"')
        author_works_rows = _fetch_articles_by_ids(conn, author_ids)
        
        articles_with_citations = _add_dynamic_citations_to_results(conn, author_works_rows, start_year, end_year)
        
        # Filter again by author name for accuracy
        final_articles = [
            work for work in articles_with_citations 
            if author_name in work.get('authors', '')
        ]

        final_articles.sort(key=lambda x: x['citedby_count'], reverse=True)
    finally:
        conn.close()
        
    return jsonify({"success": True, "articles": final_articles})

@app.route('/api/get-author-citation-stats', methods=['POST'])
def get_author_citation_stats_endpoint():
    data = request.get_json()
    author_name, start_year, end_year = data.get('author_id'), int(data.get('start_year', 2020)), int(data.get('end_year', datetime.now().year))
    if not author_name: raise ApiException("未提供作者信息。")
    conn = get_db_connection()
    try:
        author_ids = _fts_query_rowids(conn, "Authors", f'"{author_name}"')
        target_rows = _fetch_articles_by_ids(conn, author_ids)
        citing_articles_rows = _get_citing_articles_for_targets(conn, target_rows, start_year, end_year)
        
        stats = {"外语C刊": 0, "外语核心期刊": 0, "外语C集刊": 0, "total": 0}
        for row in citing_articles_rows:
            stats['total'] += 1
            journal_type = get_journal_type(row['Journal_Name'], row['Year'])
            if journal_type == '外语C刊':
                stats['外语C刊'] += 1
            elif journal_type == '外语核心期刊':
                stats['外语核心期刊'] += 1
            elif journal_type == '外语C集刊':
                stats['外语C集刊'] += 1
    finally: conn.close()
    return jsonify({"success": True, "stats": stats})

@app.route('/api/get-author-yearly-stats', methods=['POST'])
def get_author_yearly_stats_endpoint():
    data = request.get_json()
    author_name = data.get('author_id')
    if not author_name: raise ApiException("未提供作者ID。")
    conn = get_db_connection()
    try:
        author_ids = _fts_query_rowids(conn, "Authors", f'"{author_name}"')
        target_rows = _fetch_articles_by_ids(conn, author_ids)
        pub_stats = defaultdict(int)
        for row in target_rows:
            year = row['Year']
            if year is not None:
                pub_stats[year] += 1
        
        citation_stats = defaultdict(int)
        if target_rows:
            citing_articles = _get_citing_articles_for_targets(conn, target_rows, 1900, 9999)
            for article in citing_articles:
                citation_stats[article['Year']] += 1
            
        all_years_with_data = set(pub_stats.keys()) | set(citation_stats.keys())
        
        if not all_years_with_data:
            max_year = datetime.now().year
            min_year = max_year - 10
            years_range = list(range(min_year, max_year + 1))
        else:
            min_year = min(y for y in all_years_with_data if y is not None)
            max_year = max(y for y in all_years_with_data if y is not None)
            years_range = list(range(min_year, max_year + 1))

        chart_data = {
            'labels': years_range,
            'publications': [pub_stats.get(year, 0) for year in years_range],
            'citations': [citation_stats.get(year, 0) for year in years_range]
        }
    finally: 
        conn.close()
    return jsonify({"success": True, "chart_data": chart_data})


@app.route('/api/search-author-citing', methods=['POST'])
def search_author_citing_endpoint():
    data = request.get_json()
    citing_authors_str = data.get('citing_author', '').strip()
    target_journals_str = data.get('target_journals', '').strip()
    start_year, end_year = data.get('start_year'), data.get('end_year')

    citing_authors_list = [a.strip() for a in citing_authors_str.split(',') if a.strip()]
    target_journals_list = [j.strip() for j in target_journals_str.split(',') if j.strip()]
    
    log_search_activity('查作者引用', author=citing_authors_str, target_journals=target_journals_str, years=f'{start_year}-{end_year}')

    if not citing_authors_list or not target_journals_list:
        raise ApiException("作者和目标期刊均为必填项（均支持多个，用英文逗号分隔）。")

    conn = get_db_connection()
    try:
        author_fts_expr = ' OR '.join([f'"{author}"' for author in citing_authors_list])
        author_article_ids = _fts_query_rowids(conn, "Authors", author_fts_expr)

        if not author_article_ids:
            return jsonify({"success": True, "count": 0, "articles": [], "chart_data": {}})

        rows = _fetch_articles_by_ids(conn, author_article_ids, start_year=start_year, end_year=end_year)
        if not rows:
            return jsonify({"success": True, "count": 0, "articles": [], "chart_data": {}})

        temp_articles = _db_rows_to_articles_json(rows)
        articles_with_match_info = []
        citation_counts = defaultdict(int)

        if has_article_journal_links_table():
            target_norm_to_display = _normalize_target_journals(target_journals_list)
            target_norms = set(target_norm_to_display.keys())
            article_ids = [row["Article_ID"] for row in rows]
            target_links_map = (
                _fetch_journal_links(conn, article_ids, target_norms)
                if target_norms else defaultdict(set)
            )

            for article in temp_articles:
                found_author = next((author for author in citing_authors_list if author in article.get('authors', '')), None)
                if not found_author:
                    continue

                article_id = int(str(article.get("id", "")).replace("local:", "") or 0)
                matched_journals = []
                matched_norms = target_links_map.get(article_id, set())

                for norm_name in sorted(matched_norms):
                    display_name = target_norm_to_display.get(norm_name, norm_name)
                    matched_journals.append(display_name)
                    citation_counts[(found_author, display_name)] += 1

                if matched_journals:
                    article["matched_for"] = f"引用了: {', '.join(sorted(set(matched_journals)))}"
                    articles_with_match_info.append(article)
        else:
            ref_article_ids = set()
            journal_fts_expr = build_journal_fts_query(target_journals_list)
            target_ref_ids = _fts_query_rowids(conn, "Reference", journal_fts_expr)
            ref_article_ids.update(target_ref_ids)

            if not ref_article_ids:
                return jsonify({"success": True, "count": 0, "articles": [], "chart_data": {}})

            final_article_ids = author_article_ids.intersection(ref_article_ids)
            if not final_article_ids:
                return jsonify({"success": True, "count": 0, "articles": [], "chart_data": {}})

            rows = _fetch_articles_by_ids(conn, final_article_ids, start_year=start_year, end_year=end_year)
            temp_articles = _db_rows_to_articles_json(rows)

            for article in temp_articles:
                found_author = next((author for author in citing_authors_list if author in article.get('authors', '')), None)
                if not found_author:
                    continue

                matched_journals = []
                ref_text = article.get('reference', '')

                if ref_text:
                    for journal in target_journals_list:
                        base_names = get_journal_name_variants(journal)
                        for base_name in base_names:
                            if base_name in ref_text:
                                matched_journals.append(journal)
                                citation_counts[(found_author, journal)] += 1
                                break

                if matched_journals:
                    article['matched_for'] = f"引用了: {', '.join(sorted(list(set(matched_journals))))}"
                    articles_with_match_info.append(article)
        
        chart_data = {}
        if citation_counts:
            source_nodes = set(citing_authors_list)
            target_nodes = {journal for _, journal in citation_counts.keys()}
            all_node_names = source_nodes.union(target_nodes)
            sankey_nodes = [{'name': name} for name in all_node_names]
            sankey_links = [{'source': author, 'target': journal, 'value': count} for (author, journal), count in citation_counts.items() if count > 0]
            chart_data = {'sankey_nodes': sankey_nodes, 'sankey_links': sankey_links}

    finally:
        conn.close()
    
    return jsonify({"success": True, "count": len(articles_with_match_info), "articles": articles_with_match_info, "chart_data": chart_data})


# ── 作者关系网络 ──────────────────────────────────────────────────────────
import re as _re
_REF_AUTHOR_RE = _re.compile(r'\[(?:J|M|D|C|G|N|R|S|EB/OL|EB)\]\.\s*([^.\n]{1,60}?)\.')

def _parse_ref_authors(reference_text):
    """从参考文献字符串中提取作者姓名列表（中文短名）"""
    if not reference_text:
        return []
    result = []
    for match in _REF_AUTHOR_RE.finditer(reference_text):
        author_str = match.group(1).strip()
        for a in author_str.split(';'):
            a = a.strip()
            if a and 1 < len(a) <= 10 and not _re.search(r'[【】\[\]()（）《》，。、/\\]', a):
                result.append(a)
    return result

@app.route('/api/author-network', methods=['POST'])
def author_network_endpoint():
    data = request.get_json()
    author_name = data.get('author_name', '').strip()
    start_year  = int(data.get('start_year', 2000))
    end_year    = int(data.get('end_year', datetime.now().year))
    max_coauthors = int(data.get('max_coauthors', 20))
    max_cited     = int(data.get('max_cited', 25))
    max_citing    = int(data.get('max_citing', 20))
    log_search_activity('查作者关系', query=author_name)
    if not author_name:
        raise ApiException("请输入作者姓名。")
    conn = get_db_connection()
    try:
        # 1. 获取该作者的所有文章
        author_ids = _fts_query_rowids(conn, "Authors", f'"{author_name}"')
        rows = _fetch_articles_by_ids(conn, author_ids, start_year=start_year, end_year=end_year)

        if not rows:
            conn.close()
            return jsonify({"success": True, "nodes": [], "links": [], "categories": [], "stats": {"total_articles": 0}})

        # 2. 合著关系
        coauthor_count = defaultdict(int)
        for row in rows:
            if row['Authors']:
                for a in row['Authors'].split(';'):
                    a = a.strip()
                    if a and a != author_name:
                        coauthor_count[a] += 1

        # 3. 被引作者（我引用了谁）— 从参考文献解析
        cited_count = defaultdict(int)
        for row in rows:
            for a in _parse_ref_authors(row['Reference'] or ''):
                if a != author_name:
                    cited_count[a] += 1

        # 4. 引用我的作者（谁引用了我）
        citing_ids = _fts_query_rowids(conn, "Reference", f'"{author_name}"')
        citing_rows = _fetch_articles_by_ids(conn, citing_ids, start_year=start_year, end_year=end_year)
        citing_count = defaultdict(int)
        for row in citing_rows:
            if row['Authors']:
                for a in row['Authors'].split(';'):
                    a = a.strip()
                    if a and a != author_name:
                        citing_count[a] += 1

        # 5. 筛选 top-N，避免节点过多
        top_coauthors = sorted(coauthor_count.items(), key=lambda x: -x[1])[:max_coauthors]
        coauthor_set  = {a for a, _ in top_coauthors}

        top_cited = [(a, c) for a, c in sorted(cited_count.items(), key=lambda x: -x[1])
                     if a not in coauthor_set][:max_cited]
        cited_set = {a for a, _ in top_cited}

        top_citing = [(a, c) for a, c in sorted(citing_count.items(), key=lambda x: -x[1])
                      if a not in coauthor_set and a not in cited_set][:max_citing]

        # 6. 构造图节点和连线
        nodes = [{"id": author_name, "name": author_name, "category": 0, "symbolSize": 46,
                  "label": {"show": True, "fontWeight": "bold"}}]
        links = []

        max_co = max((c for _, c in top_coauthors), default=1)
        for author, count in top_coauthors:
            size = 14 + round((count / max_co) * 20)
            nodes.append({"id": author, "name": author, "category": 1, "symbolSize": size,
                          "value": count})
            links.append({"source": author_name, "target": author, "value": count,
                          "lineStyle": {"width": min(1 + count, 4)}})

        max_cd = max((c for _, c in top_cited), default=1)
        for author, count in top_cited:
            size = 10 + round((count / max_cd) * 14)
            nodes.append({"id": author, "name": author, "category": 2, "symbolSize": size,
                          "value": count})
            links.append({"source": author_name, "target": author, "value": count,
                          "lineStyle": {"width": 1}})

        max_ci = max((c for _, c in top_citing), default=1)
        for author, count in top_citing:
            size = 10 + round((count / max_ci) * 14)
            nodes.append({"id": author, "name": author, "category": 3, "symbolSize": size,
                          "value": count})
            links.append({"source": author, "target": author_name, "value": count,
                          "lineStyle": {"width": 1}})

        categories = [
            {"name": "查询作者"},
            {"name": "合著关系"},
            {"name": "引出关系（我引用）"},
            {"name": "引入关系（引用我）"},
        ]
        stats = {
            "total_articles": len(rows),
            "coauthors": len(top_coauthors),
            "cited_out": len(top_cited),
            "citing_in": len(top_citing),
        }
        return jsonify({"success": True, "nodes": nodes, "links": links,
                        "categories": categories, "stats": stats,
                        "center": author_name})
    finally:
        conn.close()


@app.route('/api/search-journal-citations', methods=['POST'])
def search_journal_citations_endpoint():
    data = request.get_json()
    source_journals = data.get('source_journals', '')
    target_journals = data.get('target_journals', '')
    start_year, end_year = data.get('start_year'), data.get('end_year')
    
    log_search_activity('查期刊互引', source_journals=source_journals, target_journals=target_journals, years=f'{start_year}-{end_year}')
    
    if not all([source_journals, target_journals]):
        raise ApiException("源期刊和目标期刊均为必填项。")
        
    source_journals_list = [j.strip() for j in source_journals.split(',') if j.strip()]
    target_journals_list = [j.strip() for j in target_journals.split(',') if j.strip()]
    
    conn = get_db_connection()
    try:
        actual_source_journals = find_matching_journals_in_db(conn, source_journals_list)
        if not actual_source_journals:
            return jsonify({"success": True, "count": 0, "articles": [], "chart_data": {}})

        placeholders = ','.join(['?'] * len(actual_source_journals))
        source_query = f"""
            SELECT * FROM articles
            WHERE Journal_Name IN ({placeholders})
            AND Year >= ? AND Year <= ?
        """
        source_params = actual_source_journals + [int(start_year), int(end_year)]
        source_rows = conn.execute(source_query, tuple(source_params)).fetchall()
        if not source_rows:
            return jsonify({"success": True, "count": 0, "articles": [], "chart_data": {}})

        all_articles = []
        heatmap_data = defaultdict(lambda: defaultdict(int))

        if has_article_journal_links_table():
            target_norm_to_display = _normalize_target_journals(target_journals_list)
            target_norms = set(target_norm_to_display.keys())
            source_article_ids = [row["Article_ID"] for row in source_rows]
            links_map = _fetch_journal_links(conn, source_article_ids, target_norms)

            for row in source_rows:
                article_id = row["Article_ID"]
                matched_norms = links_map.get(article_id, set())
                if not matched_norms:
                    continue

                matched_targets = [
                    target_norm_to_display.get(norm_name, norm_name)
                    for norm_name in sorted(matched_norms)
                ]
                article_json = _db_rows_to_articles_json([row])[0]
                article_json['matched_for'] = f"引用了: {', '.join(sorted(set(matched_targets)))}"
                all_articles.append(article_json)

                source_normalized = normalize_journal_name(row['Journal_Name'])
                year = int(row['Year'])
                heatmap_data[source_normalized][year] += 1
        else:
            target_journals_fts = build_journal_fts_query(target_journals_list)
            target_ref_ids = _fts_query_rowids(conn, "Reference", target_journals_fts)
            source_article_ids = {row["Article_ID"] for row in source_rows}
            matched_article_ids = source_article_ids.intersection(target_ref_ids)
            all_articles_rows = _fetch_articles_by_ids(
                conn,
                matched_article_ids,
                start_year=start_year,
                end_year=end_year,
            )

            for row in all_articles_rows:
                article = dict(row)
                ref_text = article.get('Reference', '')
                matched_targets = []

                if ref_text:
                    for target_journal in target_journals_list:
                        base_names = get_journal_name_variants(target_journal)
                        for base_name in base_names:
                            if base_name in ref_text:
                                matched_targets.append(target_journal)
                                break

                if matched_targets:
                    article_json = _db_rows_to_articles_json([article])[0]
                    article_json['matched_for'] = f"引用了: {', '.join(sorted(list(set(matched_targets))))}"
                    all_articles.append(article_json)

                    source_normalized = normalize_journal_name(article['Journal_Name'])
                    year = int(article['Year'])
                    heatmap_data[source_normalized][year] += 1
        
        chart_data = {}
        if heatmap_data:
            years_range = list(range(int(start_year), int(end_year) + 1))
            heatmap_matrix = []
            source_journals_with_data = []
            
            normalized_user_inputs = sorted(list(set(normalize_journal_name(j) for j in source_journals_list)))
            
            for norm_journal in normalized_user_inputs:
                if norm_journal in heatmap_data and sum(heatmap_data[norm_journal].values()) > 0:
                    source_journals_with_data.append(norm_journal)
                    row = [heatmap_data[norm_journal].get(year, 0) for year in years_range]
                    heatmap_matrix.append(row)
            
            if heatmap_matrix:
                target_display = ', '.join(target_journals_list)
                chart_data = {
                    'heatmap_data': heatmap_matrix,
                    'x_labels': [str(y) for y in years_range],
                    'y_labels': source_journals_with_data,
                    'target_journal': target_display
                }

    finally:
        conn.close()
    
    return jsonify({"success": True, "count": len(all_articles), "articles": all_articles, "chart_data": chart_data})

@app.route('/api/find-article', methods=['POST'])
def find_article_endpoint():
    data = request.get_json()
    identifier = data.get('identifier', '').strip()
    log_search_activity('查文章被引', identifier=identifier)
    if not identifier: raise ApiException("请输入文章标识。")
    conn = get_db_connection()

    cursor = conn.execute("SELECT * FROM articles WHERE DOI = ? OR Article_ID = ?", (str(identifier), str(identifier)))
    articles = list(cursor.fetchall())
    
    if not articles:
        title_ids = _fts_query_rowids(conn, "Title", f'"{identifier}"')
        articles.extend(_fetch_articles_by_ids(conn, title_ids))

    articles_json = _add_dynamic_citations_to_results(conn, articles, start_year=2000, end_year=datetime.now().year)
    conn.close()
    return jsonify({"success": True, "articles": articles_json})

@app.route('/api/analyze-article-citations', methods=['POST'])
def analyze_article_citations_endpoint():
    data = request.get_json()
    target_article_id, start_year, end_year = data.get('article_id'), data.get('start_year'), data.get('end_year')
    conn = get_db_connection()
    try:
        clean_id = str(target_article_id).replace('local:', '')
        cursor = conn.execute("SELECT * FROM articles WHERE Article_ID = ?", (clean_id,))
        row = cursor.fetchone()
        if not row: raise ApiException("无法检索到文章详情。", 404)
        target_article_details = _db_rows_to_articles_json([row])[0]
        citing_articles_rows = _get_citing_articles_for_targets(conn, [row], start_year, end_year)
        stats = {"外语C刊": 0, "外语核心期刊": 0, "外语C集刊": 0, "total": 0}
        for r in citing_articles_rows:
            stats['total'] += 1
            jt = get_journal_type(r['Journal_Name'], r['Year'])
            if jt == '外语C刊':
                stats['外语C刊'] += 1
            elif jt == '外语核心期刊':
                stats['外语核心期刊'] += 1
            elif jt == '外语C集刊':
                stats['外语C集刊'] += 1
    finally: conn.close()
    return jsonify({"success": True, "details": target_article_details, "stats": stats})

@app.route('/api/get-citing-articles', methods=['POST'])
def get_citing_articles_endpoint():
    data = request.get_json()
    target_id, target_type = data.get('target_id'), data.get('target_type')
    citation_type, start_year, end_year = data.get('citation_type'), data.get('start_year'), data.get('end_year')
    if not all([target_id, target_type, citation_type, start_year, end_year]): raise ApiException("缺少必要的参数。")
    conn = get_db_connection()
    try:
        target_rows = []
        if target_type == 'author':
            author_ids = _fts_query_rowids(conn, "Authors", f'"{target_id}"')
            target_rows = _fetch_articles_by_ids(conn, author_ids)
        elif target_type == 'article':
            clean_id = str(target_id).replace('local:', '')
            cursor = conn.execute("SELECT Article_ID, Title, Authors, Year, Journal_Name FROM articles WHERE Article_ID = ?", (clean_id,))
            row = cursor.fetchone()
            if row and row['Title']:
                target_rows = [row]
        
        all_citing_rows = _get_citing_articles_for_targets(conn, target_rows, start_year, end_year)
        
        journal_type_map = {'CSSCI': '外语C刊', 'Core': '外语核心期刊', 'Collective': '外语C集刊'}
        if citation_type == 'Total':
            filtered_articles = _db_rows_to_articles_json(all_citing_rows)
        else:
            target_journal_type = journal_type_map.get(citation_type)
            if target_journal_type:
                filtered_rows = [row for row in all_citing_rows if get_journal_type(row['Journal_Name'], row['Year']) == target_journal_type]
            else:
                filtered_rows = []
            filtered_articles = _db_rows_to_articles_json(filtered_rows)
    finally: conn.close()
    return jsonify({"success": True, "articles": filtered_articles})


# --- Excel导出功能 ---
@app.route('/api/export-data', methods=['POST'])
def export_data_endpoint():
    data = request.get_json()
    articles, filename = data.get('articles', []), data.get('filename', 'export.xlsx')
    if not articles:
        return jsonify({"success": False, "error": "没有可导出的数据。"}), 400
    df = pd.DataFrame(articles)
    
    export_cols = {
        'title': '标题', 
        'authors': '作者', 
        'publicationName': '期刊', 
        'coverDate': '年份', 
        'citedby_count': '被引数',
        'matched_for': '引用的目标期刊'
    }
    
    df_export = pd.DataFrame()
    for col_key, col_name in export_cols.items():
        if col_key in df:
            df_export[col_name] = df[col_key]
        else:
            df_export[col_name] = ''
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in worksheet["1:1"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for col_cells in worksheet.columns:
            max_length = 0
            column = col_cells[0].column_letter
            for cell in col_cells:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = min(adjusted_width, 60)
            
    output.seek(0)
    
    # FIX: 对含有非ASCII字符的文件名进行编码
    ascii_filename = 'export.xlsx'
    encoded_filename = quote(filename)
    headers = {
        "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}"
    }
    
    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@app.route('/api/export-heatmap-data', methods=['POST'])
def export_heatmap_data_endpoint():
    """导出热力图数据为Excel，并进行美化和总计"""
    data = request.get_json()
    heatmap_data = data.get('data', [])
    filename = data.get('filename', 'heatmap_export.xlsx')
    target_journal = data.get('target_journal', '目标期刊')
    
    if not heatmap_data or len(heatmap_data) < 2:
        return jsonify({"success": False, "error": "没有可导出的数据。"}), 400
    
    # 1. 创建DataFrame并剔除前端传入的多余表头行
    df = pd.DataFrame(heatmap_data)
    df = df.iloc[1:].reset_index(drop=True)
    
    # 2. 重命名列
    df = df.rename(columns={'期刊名称': '源期刊'})
    year_columns_map = {col: col.replace('year_', '') for col in df.columns if col.startswith('year_')}
    df = df.rename(columns=year_columns_map)
    
    year_columns = list(year_columns_map.values())
    df[year_columns] = df[year_columns].apply(pd.to_numeric)
    
    # 3. 增加“总计”行和列
    df['总计'] = df[year_columns].sum(axis=1)
    total_row = df[year_columns + ['总计']].sum()
    total_row['源期刊'] = '总计'
    df = pd.concat([df, pd.DataFrame(total_row).T], ignore_index=True)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='互引数据')
        worksheet = writer.sheets['互引数据']
        
        # --- 4. 美化表格 ---
        
        # 定义样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bold_font = Font(bold=True)
        zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        grand_total_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        # 插入并美化大标题
        worksheet.insert_rows(1)
        last_col_letter = get_column_letter(worksheet.max_column)
        worksheet.merge_cells(f'A1:{last_col_letter}1')
        title_cell = worksheet['A1']
        title_cell.value = f'期刊互引数据 - 引用《{target_journal}》统计'
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        worksheet.row_dimensions[1].height = 30
        
        # 美化列表头
        for cell in worksheet["2:2"]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # ✅ FIX: 调整美化逻辑，将总计行纳入斑马纹计算
        last_row = worksheet.max_row
        last_col = worksheet.max_column
        
        # 遍历所有数据行（包括总计）应用斑马纹和通用样式
        for i, row in enumerate(worksheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=last_col)):
            is_zebra_row = i % 2 == 1  # 决定是否应用斑马纹
            is_total_row = (row[0].row == last_row) # 判断是否为总计行

            for cell in row:
                cell.border = thin_border
                # 应用斑马纹
                if is_zebra_row:
                    cell.fill = zebra_fill
                # 为总计行所有单元格设置粗体
                if is_total_row:
                    cell.font = bold_font
                # 统一居中对齐
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # 为总计列（最后一列）所有单元格设置粗体
        for row in worksheet.iter_rows(min_row=3, max_row=last_row, min_col=last_col, max_col=last_col):
            row[0].font = bold_font

        # 特殊样式化右下角的总计单元格（会覆盖之前的样式）
        grand_total_cell = worksheet.cell(row=last_row, column=last_col)
        grand_total_cell.fill = grand_total_fill
        
        # 调整列宽
        for i, cell in enumerate(worksheet[2]):
            column_letter = cell.column_letter
            if i == 0:
                worksheet.column_dimensions[column_letter].width = 30
            elif i == last_col - 1:
                worksheet.column_dimensions[column_letter].width = 12
            else:
                worksheet.column_dimensions[column_letter].width = 10
            
    output.seek(0)

    # FIX: 对含有非ASCII字符的文件名进行编码
    ascii_filename = 'heatmap_export.xlsx'
    encoded_filename = quote(filename)
    headers = {
        "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}"
    }

    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

if __name__ == '__main__':
    debug_enabled = os.environ.get("FLASK_DEBUG", "0") == "1"
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5000"))
    app.run(debug=debug_enabled, host=host, port=port)
